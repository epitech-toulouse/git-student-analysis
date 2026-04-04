#!/usr/bin/env python3
"""
Génère le classeur Excel d'analyse Git.
Usage: python generate_xlsx.py report.json output.xlsx
"""
import sys, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCORE_LABELS = {3: "⭐⭐⭐ Excellent", 2: "⭐⭐ Correct", 1: "⭐ Insuffisant", 0: "⚠️ Vide/Inutile"}
FREQ_COLORS = {
    "régulière": "C6EFCE",
    "rush en fin de projet": "FFEB9C",
    "concentrée sur 1–2 jours": "FFEB9C",
    "inconnue": "F2F2F2",
}

HDR_FILL = PatternFill("solid", start_color="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
ALERT_FILL = PatternFill("solid", start_color="FFD966")
INACTIVE_FILL = PatternFill("solid", start_color="FF7070")
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

def auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

def apply_header(ws, headers):
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = BORDER

def cell_border(row):
    for cell in row:
        cell.border = BORDER

def generate(report_path: str, output_path: str):
    with open(report_path) as f:
        data = json.load(f)

    wb = Workbook()

    # ── Onglet Résumé ──────────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Résumé"

    summary = data["repo_summary"]
    ws_sum.append(["Rapport Git — Analyse pédagogique"])
    ws_sum["A1"].font = Font(bold=True, size=14, name="Arial")
    ws_sum.append([f"Période : {summary['project_start']} → {summary['project_end']}  |  Durée : {summary['project_days']} jours"])
    ws_sum.append([f"Total commits : {summary['total_commits']}  |  Auteurs (après déduplication) : {summary['authors_deduplicated']}"])
    ws_sum.append([])

    headers = ["Étudiant", "Email/Clé", "Commits", "% Commits", "+Lignes", "-Lignes",
               "Fichiers", "1er commit", "Dernier commit", "Fréquence",
               "Qualité messages (moy)", "Alertes"]
    apply_header(ws_sum, headers)

    for s in data["students"]:
        score_label = SCORE_LABELS.get(round(s["avg_message_score"]), "?")
        alerts_text = " | ".join(s["alerts"]) if s["alerts"] else ""
        row = [
            s["display_name"],
            s["canonical_key"],
            s["commit_count"],
            f"{s['commit_pct']}%",
            s["insertions"],
            s["deletions"],
            s["files_distinct"],
            s["first_commit"][:10] if s["first_commit"] else "",
            s["last_commit"][:10] if s["last_commit"] else "",
            s["frequency_label"],
            score_label,
            alerts_text,
        ]
        ws_sum.append(row)
        last_row = ws_sum[ws_sum.max_row]
        cell_border(last_row)
        # Colorier selon alertes
        if s["commit_count"] == 0:
            for cell in last_row:
                cell.fill = INACTIVE_FILL
        elif s["alerts"]:
            for cell in last_row:
                cell.fill = ALERT_FILL
        else:
            freq_color = FREQ_COLORS.get(s["frequency_label"], "F2F2F2")
            for cell in last_row[9:10]:  # colonne Fréquence seulement
                cell.fill = PatternFill("solid", start_color=freq_color)

    ws_sum.freeze_panes = "A6"
    auto_width(ws_sum)

    # ── Onglet Commits ─────────────────────────────────────────────────────────
    ws_commits = wb.create_sheet("Commits")
    apply_header(ws_commits, ["Étudiant", "Hash", "Date", "Message", "Score qualité", "+Lignes", "-Lignes"])

    for s in data["students"]:
        for c in s["commits"]:
            row = [
                s["display_name"],
                c["hash"],
                c["date"][:16].replace("T", " ") if c["date"] else "",
                c["message"],
                SCORE_LABELS.get(c["score"], "?"),
                "",  # insertions not per-commit in current extract
                "",
            ]
            ws_commits.append(row)
            last_row = ws_commits[ws_commits.max_row]
            cell_border(last_row)
            if c["score"] == 0:
                last_row[4].fill = ALERT_FILL

    ws_commits.freeze_panes = "A2"
    auto_width(ws_commits)

    # ── Onglet Alertes ─────────────────────────────────────────────────────────
    ws_alerts = wb.create_sheet("Alertes")
    apply_header(ws_alerts, ["Étudiant", "Type d'alerte", "Détail"])

    has_alerts = False
    for s in data["students"]:
        for alert in s["alerts"]:
            ws_alerts.append([s["display_name"], "⚠️ Alerte", alert])
            last_row = ws_alerts[ws_alerts.max_row]
            cell_border(last_row)
            for cell in last_row:
                cell.fill = ALERT_FILL
            has_alerts = True

    if not has_alerts:
        ws_alerts.append(["✅ Aucune alerte détectée", "", ""])

    # ── Onglet Fusions ─────────────────────────────────────────────────────────
    if data.get("merges"):
        ws_merge = wb.create_sheet("Identités fusionnées")
        apply_header(ws_merge, ["Email source", "Fusionné vers", "Nom canonique"])
        for m in data["merges"]:
            ws_merge.append([m["from"], m["to"], m["name"]])
            cell_border(ws_merge[ws_merge.max_row])
        auto_width(ws_merge)

    auto_width(ws_alerts)
    wb.save(output_path)
    print(f"Excel généré : {output_path}")

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python generate_xlsx.py <data.json> <output.xlsx>")
        sys.exit(1)

    generate(sys.argv[1], sys.argv[2])
